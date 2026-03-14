"""
Catalog Extractor Terintegrasi - Versi dengan OCR untuk Nota dari Excel
Membaca semua sheet dan mencocokkan transaksi dengan nota menggunakan OCR
"""

import sys
print(f"Python version: {sys.version}")

import pandas as pd
import json
import argparse
from pathlib import Path
from typing import Optional, Dict, List, Any, Tuple
import re
from openai import OpenAI
import os
from datetime import datetime
import numpy as np
import base64
from PIL import Image
import io
import tempfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import zipfile

# Configuration
try:
    from config import Config
    DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY", Config.DASHSCOPE_API_KEY)
    QWEN_BASE_URL = os.getenv("QWEN_BASE_URL", Config.QWEN_BASE_URL)
    QWEN_MAX_MODEL = os.getenv("QWEN_MAX_MODEL", Config.QWEN_MAX_MODEL)
    QWEN_VL_MODEL = os.getenv("QWEN_VL_MODEL", Config.QWEN_VL_MODEL)
except ImportError:
    DASHSCOPE_API_KEY = os.getenv("DASHSCOPE_API_KEY", "sk-0fc24c229d174d1b99624a49544b1c7a")
    QWEN_BASE_URL = "https://dashscope-intl.aliyuncs.com/compatible-mode/v1"
    QWEN_MAX_MODEL = "qwen-max"
    QWEN_VL_MODEL = "qwen-vl-max"


class ExcelImageExtractor:
    """Ekstrak gambar dari Excel file"""
    
    @staticmethod
    def extract_images_from_excel(excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """
        Extract all images from Excel file and map them to sheets
        Returns: Dict[sheet_name, List[image_info]]
        """
        print(f"\n[IMAGE EXTRACTOR] Extracting images from Excel: {excel_path}")
        
        images_by_sheet = {}
        
        try:
            # Method 1: Using openpyxl to extract images
            wb = load_workbook(excel_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_images = []
                
                # Check for images in the sheet
                if hasattr(ws, '_images') and ws._images:
                    for idx, img in enumerate(ws._images):
                        try:
                            # Extract image data
                            img_data = img._data()
                            
                            # Save to temp file
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                                tmp_file.write(img_data)
                                tmp_path = tmp_file.name
                            
                            # Get image position (approx row/col)
                            anchor = img.anchor
                            row = getattr(anchor, '_from', {}).row if hasattr(anchor, '_from') else 0
                            col = getattr(anchor, '_from', {}).col if hasattr(anchor, '_from') else 0
                            
                            sheet_images.append({
                                'path': tmp_path,
                                'row': row,
                                'col': col,
                                'index': idx
                            })
                            print(f"  ✓ Found image in {sheet_name} at row ~{row}, col ~{col}")
                        except Exception as e:
                            print(f"  ✗ Error extracting image from {sheet_name}: {e}")
                
                if sheet_images:
                    images_by_sheet[sheet_name] = sheet_images
            
            # Method 2: If no images found with openpyxl, try extracting from Excel as zip
            if not images_by_sheet:
                images_by_sheet = ExcelImageExtractor._extract_from_zip(excel_path)
            
            # Count total images
            total_images = sum(len(images) for images in images_by_sheet.values())
            print(f"\n[IMAGE EXTRACTOR] Found {total_images} images in Excel file")
            
            return images_by_sheet
            
        except Exception as e:
            print(f"  ✗ Error in image extraction: {e}")
            return {}
    
    @staticmethod
    def _extract_from_zip(excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Extract images by treating Excel as ZIP file"""
        images_by_sheet = {}
        
        try:
            with zipfile.ZipFile(excel_path, 'r') as zip_file:
                # Look for images in the media folder
                media_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]
                
                for media_file in media_files:
                    # Extract image
                    with tempfile.NamedTemporaryFile(suffix=Path(media_file).suffix, delete=False) as tmp_file:
                        tmp_file.write(zip_file.read(media_file))
                        tmp_path = tmp_file.name
                    
                    # Try to determine which sheet this image belongs to
                    # This is tricky - we'll associate with first sheet as default
                    sheet_name = "Unknown"
                    if 'drawings' in zip_file.namelist():
                        # Parse drawing relationships to map images to sheets
                        pass
                    
                    if sheet_name not in images_by_sheet:
                        images_by_sheet[sheet_name] = []
                    
                    images_by_sheet[sheet_name].append({
                        'path': tmp_path,
                        'filename': media_file,
                        'index': len(images_by_sheet[sheet_name])
                    })
                    
                    print(f"  ✓ Found embedded image: {media_file}")
        
        except Exception as e:
            print(f"  ✗ Error in ZIP extraction: {e}")
        
        return images_by_sheet


class ReceiptOCRProcessor:
    """OCR Processor untuk membaca nota menggunakan Qwen VL"""
    
    def __init__(self, api_key: str = None, base_url: str = None):
        """Initialize dengan OpenAI client untuk Qwen VL"""
        self.api_key = api_key or DASHSCOPE_API_KEY
        self.base_url = base_url or QWEN_BASE_URL
        self.vl_model = QWEN_VL_MODEL
        
        self.client = OpenAI(
            api_key=self.api_key,
            base_url=self.base_url
        )
    
    def encode_image(self, image_path: str) -> str:
        """Encode gambar ke base64"""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    
    def process_receipt_image(self, image_path: str, expected_date: str = None) -> List[Dict[str, Any]]:
        """
        Process receipt image using Qwen VL to extract items and prices
        """
        print(f"\n  [OCR] Processing receipt image: {image_path}")
        if expected_date:
            print(f"  [OCR] Expected date: {expected_date}")
        
        try:
            # Encode image
            base64_image = self.encode_image(image_path)
            
            prompt = """Analyze this receipt/invoice image CAREFULLY. Extract EVERY item with its price.

            This is a detailed receipt showing multiple purchases. You MUST extract ALL items.

            Return a JSON array with this EXACT structure:
            [
                {
                    "item_name": "exact name of the item as written",
                    "item_price": 15000,
                    "quantity": 1,
                    "total_price": 15000
                }
            ]

            RULES:
            1. Extract EVERY line item from the receipt - do not miss any
            2. If an item has multiple units (e.g., "BBM KLX 2x"), record quantity=2 and item_price=unit price
            3. If total price is given but not unit price, use total_price and set quantity=1
            4. Common items to look for:
            - BBM KLX, BBM Vixion, BBM Motor, BBM Mesin
            - Galon (air minum)
            - Tali Rafia
            - Sapu, Canebo, Tisu Mobil
            - Entertaint (makan/minum)
            - Biaya Tarik Tunai
            - Materai
            - Colokan Listrik, Terminal Listrik
            - Ban Dalam, Tambal Ban

            Return ONLY the JSON array, no other text."""

            response = self.client.chat.completions.create(
                model=self.vl_model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": f"data:image/png;base64,{base64_image}"
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ],
                temperature=0.1,
                max_tokens=2000
            )
            
            result_text = response.choices[0].message.content.strip()
            print(f"  [OCR] Raw response: {result_text[:200]}...")
            
            # Extract JSON array
            start_idx = result_text.find('[')
            end_idx = result_text.rfind(']') + 1
            
            if start_idx >= 0 and end_idx > start_idx:
                json_str = result_text[start_idx:end_idx]
                items = json.loads(json_str)
                print(f"  [OCR] Successfully extracted {len(items)} items from receipt")
                
                # Validate items
                valid_items = []
                for item in items:
                    if item.get('item_name') and (item.get('item_price') or item.get('total_price')):
                        valid_items.append(item)
                        print(f"    - {item.get('item_name')}: Rp {item.get('item_price') or item.get('total_price')}")
                
                return valid_items
            else:
                print(f"  [OCR] No JSON array found in response")
                return []
                
        except Exception as e:
            print(f"  [OCR] Error processing receipt: {str(e)}")
            return []
    
    def extract_receipt_date(self, image_path: str) -> Optional[str]:
        """Extract date from receipt image"""
        try:
            base64_image = self.encode_image(image_path)
            
            prompt = """What is the date on this receipt? Return only the date in YYYY-MM-DD format. If no date found, return null."""
            
            response = self.client.chat.completions.create(
                model=self.vl_model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image_url",
                                "image_url": f"data:image/png;base64,{base64_image}"
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ],
                temperature=0.1
            )
            
            date_str = response.choices[0].message.content.strip()
            if date_str and date_str != 'null':
                return date_str
            return None
            
        except Exception as e:
            print(f"  [OCR] Error extracting date: {e}")
            return None


class MultiSheetExcelExtractor:
    """Excel Extractor untuk membaca semua sheet"""
    
    def __init__(self):
        """Initialize extractor"""
        self.transaction_patterns = [
            r'BBM\s+(\w+)',
            r'Tali\s+Rafia',
            r'Galon',
            r'Entertaint',
            r'Biaya\s+Tarik\s+Tunai',
            r'Materai',
            r'Listrik',
            r'Ban\s+Dalam',
            r'Colokan\s+Listrik',
        ]
    
    def get_all_sheets(self, excel_path: str) -> List[str]:
        """Get all sheet names from Excel file"""
        excel_file = pd.ExcelFile(excel_path)
        return excel_file.sheet_names
    
    def is_monitoring_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a monitoring sheet (not lampiran)"""
        return ('monitoring' in sheet_name.lower() or 
                any(month in sheet_name.lower() for month in ['januari', 'februari', 'maret', 'april', 
                                                              'mei', 'juni', 'juli', 'agustus', 
                                                              'september', 'oktober', 'november', 'desember'])) \
               and 'lampiran' not in sheet_name.lower() and 'nota' not in sheet_name.lower()
    
    def is_lampiran_sheet(self, sheet_name: str) -> bool:
        """Check if sheet is a lampiran sheet"""
        return 'lampiran' in sheet_name.lower() or 'nota' in sheet_name.lower()
    
    def extract_monitoring_data(self, excel_path: str, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract transactions from monitoring sheet"""
        print(f"\n  [EXTRACT] Processing monitoring sheet: {sheet_name}")
        
        excel_file = pd.ExcelFile(excel_path)
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        
        transactions = []
        current_date = None
        current_page = 1
        
        # Track rows to skip (like "Sisa Saldo" rows)
        skip_keywords = ['sisa saldo', 'saldo', 'rp', 'total']
        
        for idx, row in df.iterrows():
            # Skip early rows with headers
            if idx < 5:
                continue
                
            date_val = None
            operasional_val = None
            keterangan_val = None
            row_text = ""
            
            # First pass: collect all text and identify date
            for col_idx, val in enumerate(row):
                if pd.isna(val):
                    continue
                
                str_val = str(val).strip()
                row_text += " " + str_val.lower()
                
                # Check for date
                if isinstance(val, datetime):
                    date_val = val.strftime('%Y-%m-%d')
                    current_date = date_val
                    current_page += 1
                elif isinstance(val, str) and re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', val):
                    try:
                        parsed_date = pd.to_datetime(val, dayfirst=True, errors='coerce')
                        if pd.notna(parsed_date):
                            date_val = parsed_date.strftime('%Y-%m-%d')
                            current_date = date_val
                            current_page += 1
                    except:
                        pass
            
            # Skip if this is a "Sisa Saldo" row
            if any(skip in row_text for skip in skip_keywords) and len(row_text) < 30:
                continue
            
            # Second pass: extract operasional and keterangan
            for col_idx, val in enumerate(row):
                if pd.isna(val):
                    continue
                
                # Check for amount (numeric)
                if isinstance(val, (int, float, np.number)) and val != 0:
                    # This might be an amount - check if it's in OPERASIONAL column range
                    if col_idx in [4, 5, 6]:  # Columns that might contain OPERASIONAL
                        operasional_val = float(val)
                
                # Check for keterangan (text) - longer than 5 chars and not just numbers
                if isinstance(val, str) and len(val) > 5 and not re.match(r'^\d+$', val):
                    # Skip if it's just "Sisa Saldo" or similar
                    if not any(skip in val.lower() for skip in ['sisa saldo', 'rp', 'total']):
                        keterangan_val = val
            
            # If we have both date and keterangan, create transaction
            if keterangan_val and current_date:
                transaction = {
                    "tanggal": current_date,
                    "operasional": operasional_val,
                    "keterangan": keterangan_val,
                    "_sheet": sheet_name,
                    "_page": current_page,
                    "_row": idx + 1
                }
                transactions.append(transaction)
                print(f"    ✓ Found: {keterangan_val[:50]}... (Rp {operasional_val})" if operasional_val else f"    ✓ Found: {keterangan_val[:50]}... (no price)")
        
        print(f"  [EXTRACT] Found {len(transactions)} transactions in {sheet_name}")
        return transactions
    
    def extract_lampiran_data(self, excel_path: str, sheet_name: str, images: List[Dict] = None) -> Dict[str, Any]:
        """Extract receipt data from lampiran sheet"""
        print(f"\n  [EXTRACT] Processing lampiran sheet: {sheet_name}")
        
        excel_file = pd.ExcelFile(excel_path)
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        
        receipt_data = {
            "sheet_name": sheet_name,
            "dates": [],
            "images": images or []
        }
        
        # Find dates in the sheet (these indicate receipt sections)
        for idx, row in df.iterrows():
            for col_idx, val in enumerate(row):
                if pd.isna(val):
                    continue
                
                # Look for dates
                if isinstance(val, datetime):
                    date_str = val.strftime('%Y-%m-%d')
                    receipt_data["dates"].append({
                        "date": date_str,
                        "row": idx + 1,
                        "col": col_idx + 1
                    })
                    print(f"    ✓ Found receipt date: {date_str} at row {idx+1}")
                elif isinstance(val, str):
                    # Try to parse as date
                    date_match = re.search(r'(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})', val)
                    if date_match:
                        try:
                            parsed_date = pd.to_datetime(val, dayfirst=True, errors='coerce')
                            if pd.notna(parsed_date):
                                date_str = parsed_date.strftime('%Y-%m-%d')
                                receipt_data["dates"].append({
                                    "date": date_str,
                                    "row": idx + 1,
                                    "col": col_idx + 1
                                })
                                print(f"    ✓ Found receipt date: {date_str} at row {idx+1}")
                        except:
                            pass
        
        print(f"  [EXTRACT] Found {len(receipt_data['dates'])} receipt dates and {len(receipt_data['images'])} images in {sheet_name}")
        return receipt_data


class IntegratedCatalogExtractor:
    """Integrated extractor untuk semua sheet dengan OCR"""
    
    def __init__(self, api_key: str = None):
        """Initialize extractors"""
        self.excel_extractor = MultiSheetExcelExtractor()
        self.llm_processor = LLMSemanticProcessor(api_key=api_key)
        self.ocr_processor = ReceiptOCRProcessor(api_key=api_key)
        self.image_extractor = ExcelImageExtractor()
        
        # Storage for all data
        self.all_transactions = []
        self.all_receipt_data = []  # List of receipt data per sheet
        self.receipt_items_cache = {}  # date -> list of items from OCR
    
    def process_all_sheets(self, excel_path: str, output_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Complete workflow: Process all sheets, extract images, match with OCR
        """
        print(f"\n{'='*70}")
        print(f"INTEGRATED CATALOG EXTRACTOR - MULTI SHEET WITH OCR")
        print(f"{'='*70}\n")
        
        try:
            # Step 1: Extract all images from Excel
            print("\n[STEP 1] Extracting images from Excel...")
            all_images = self.image_extractor.extract_images_from_excel(excel_path)
            
            # Step 2: Get all sheets
            all_sheets = self.excel_extractor.get_all_sheets(excel_path)
            print(f"\n[STEP 2] Found {len(all_sheets)} sheets in Excel file:")
            for sheet in all_sheets:
                print(f"  - {sheet}")
            
            # Step 3: Separate and process sheets
            monitoring_sheets = []
            lampiran_sheets = []
            
            for sheet in all_sheets:
                if self.excel_extractor.is_monitoring_sheet(sheet):
                    monitoring_sheets.append(sheet)
                elif self.excel_extractor.is_lampiran_sheet(sheet):
                    lampiran_sheets.append(sheet)
            
            print(f"\n[STEP 3] Processing sheets...")
            print(f"  Monitoring sheets: {len(monitoring_sheets)}")
            print(f"  Lampiran sheets: {len(lampiran_sheets)}")
            
            # Process monitoring sheets
            print(f"\n[STEP 4] Extracting transactions from monitoring sheets...")
            for sheet in monitoring_sheets:
                transactions = self.excel_extractor.extract_monitoring_data(excel_path, sheet)
                self.all_transactions.extend(transactions)
            
            # Process lampiran sheets with images
            print(f"\n[STEP 5] Processing lampiran sheets and images...")
            for sheet in lampiran_sheets:
                sheet_images = all_images.get(sheet, [])
                receipt_data = self.excel_extractor.extract_lampiran_data(excel_path, sheet, sheet_images)
                self.all_receipt_data.append(receipt_data)
            
            print(f"\n[STEP 6] Processing receipts with OCR...")
            self._process_all_receipts()
            
            print(f"\n[STEP 7] Matching transactions with receipts...")
            matched_items = self._match_transactions_with_receipts()
            
            # Compile statistics
            print(f"\n[STEP 8] Compiling statistics...")
            items_with_price = sum(1 for item in matched_items if item.get('harga_barang') is not None)
            items_from_receipt = sum(1 for item in matched_items if item.get('_matched_with_receipt', False))
            
            result = {
                "total_items": len(matched_items),
                "stats": {
                    "total_items": len(matched_items),
                    "items_with_price": items_with_price,
                    "items_from_receipt": items_from_receipt,
                    "coverage_percent": (items_with_price / len(matched_items) * 100) if matched_items else 0,
                    "receipt_match_rate": (items_from_receipt / len(matched_items) * 100) if matched_items else 0
                },
                "sheets_processed": {
                    "monitoring": monitoring_sheets,
                    "lampiran": lampiran_sheets
                },
                "receipts_processed": len(self.receipt_items_cache),
                "items": matched_items
            }
            
            # Save result
            print(f"\n[STEP 9] Saving results...")
            output_file = self._save_result(result, excel_path, output_path)
            
            # Print summary
            self._print_summary(result, output_file)
            
            return result
            
        except Exception as e:
            print(f"\n✗ Error: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _save_result(self, result: Dict[str, Any], excel_path: str, output_path: Optional[str] = None) -> str:
        """Save result to JSON file"""
        if output_path is None:
            excel_file = Path(excel_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = excel_file.parent / f"{excel_file.stem}_catalog_result_{timestamp}.json"
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        
        print(f"  ✓ Results saved to: {output_path}")
        return str(output_path)
    
    def _print_summary(self, result: Dict[str, Any], output_file: str):
        """Print summary of results"""
        print(f"\n{'='*70}")
        print(f"✓ PROCESSING COMPLETE")
        print(f"{'='*70}")
        print(f"  Total items: {result['total_items']}")
        print(f"  Items with price: {result['stats']['items_with_price']}")
        print(f"  Items from receipt OCR: {result['stats']['items_from_receipt']}")
        print(f"  Price coverage: {result['stats']['coverage_percent']:.2f}%")
        print(f"  Receipt match rate: {result['stats']['receipt_match_rate']:.2f}%")
        print(f"  Receipts processed: {result['receipts_processed']}")
        print(f"  Output file: {output_file}")
        
        # Print sample items
        if result["items"]:
            print(f"\n📋 Sample items:")
            
            items_from_receipt = [item for item in result["items"] if item.get('_matched_with_receipt', False)]
            items_from_transaction = [item for item in result["items"] if not item.get('_matched_with_receipt', False)]
            
            if items_from_receipt:
                print(f"\n  Items from RECEIPT OCR ({len(items_from_receipt)}):")
                for i, item in enumerate(items_from_receipt[:5]):
                    price = item.get('harga_barang')
                    price_str = f"Rp {price:,.0f}" if price else "Tidak ada harga"
                    kategori = item.get('jenis_tipe_barang', 'Unknown')
                    print(f"    {i+1}. {item.get('nama_barang')[:50]} - {price_str} ({kategori})")
            
            if items_from_transaction:
                print(f"\n  Items from TRANSACTION ONLY ({len(items_from_transaction)}):")
                for i, item in enumerate(items_from_transaction[:3]):
                    price = item.get('harga_barang')
                    price_str = f"Rp {price:,.0f}" if price else "Tidak ada harga"
                    print(f"    {i+1}. {item.get('nama_barang')[:50]} - {price_str}")
    
    def _categorize_item(self, item_name: str) -> str:
        """Categorize item based on name"""
        if not item_name:
            return "Lain-lain"
        
        item_lower = item_name.lower()
        
        # Kategori Bahan Bakar
        if any(kw in item_lower for kw in ['bbm', 'bensin', 'solar', 'premium', 'pertalite', 'pertamax']):
            return "Bahan Bakar Minyak"
        
        # Kategori Air Minum
        elif any(kw in item_lower for kw in ['galon', 'air', 'aqua', 'minum']):
            return "Air Minum"
        
        # Kategori Suku Cadang
        elif any(kw in item_lower for kw in ['ban', 'ban dalam', 'ban luar', 'velg', 'rantai', 'gir', 'kampas']):
            return "Suku Cadang"
        
        # Kategori Perlengkapan Listrik
        elif any(kw in item_lower for kw in ['listrik', 'colokan', 'stop kontak', 'kabel', 'terminal', 'steker']):
            return "Perlengkapan Listrik"
        
        # Kategori Konsumsi/Entertaint
        elif any(kw in item_lower for kw in ['entertaint', 'makan', 'minum', 'kopi', 'gula', 'teh', 'snack']):
            return "Konsumsi"
        
        # Kategori Alat Tulis
        elif any(kw in item_lower for kw in ['spidol', 'pulpen', 'buku', 'kertas', 'map', 'klip', 'penjepit']):
            return "Alat Tulis"
        
        # Kategori Alat dan Bahan
        elif any(kw in item_lower for kw in ['tali', 'rafia', 'sapu', 'canebo', 'kain', 'majun', 'lap', 'amplas']):
            return "Alat dan Bahan"
        
        # Kategori Perlengkapan Kantor
        elif any(kw in item_lower for kw in ['materai', 'stempel', 'print', 'fotokopi']):
            return "Perlengkapan Kantor"
        
        # Kategori Jasa
        elif any(kw in item_lower for kw in ['tarik tunai', 'biaya', 'ongkos', 'upah', 'service', 'tambal']):
            return "Jasa"
        
        else:
            return "Lain-lain"
    
    def _split_keterangan(self, keterangan: str) -> List[str]:
        """Split keterangan into individual items"""
        if not keterangan:
            return []
        
        # Convert to lowercase for consistency
        keterangan = keterangan.lower()
        
        # Remove common prefixes/suffixes
        keterangan = re.sub(r'^(bbm|biaya|pembayaran|beli|pembelian)\s+', '', keterangan)
        keterangan = re.sub(r'\s+(dan|&)\s+', ', ', keterangan)
        
        # Common separators
        separators = [',', ';', '/', '-', 'dan', '&', r'\s+dan\s+', r'\s+&\s+']
        
        # Replace all separators with comma
        text = keterangan
        for sep in separators:
            text = re.sub(sep, ',', text)
        
        # Split and clean
        items = []
        for item in text.split(','):
            item = item.strip()
            # Filter out empty items and very short words
            if item and len(item) > 1:
                # Remove numbers and extra spaces
                item = re.sub(r'\s+', ' ', item)
                item = re.sub(r'^\d+\s*', '', item)  # Remove leading numbers
                if item and item not in ['bbm', 'rp', 'dan', '&', 'untuk', 'yang']:
                    items.append(item)
        
        # If splitting failed, return original as one item
        if not items and keterangan:
            items = [keterangan]
        
        return items
    
    def _process_all_receipts(self):
        """Process all receipt images with OCR"""
        print("\n  Processing receipt images...")
        
        for receipt_data in self.all_receipt_data:
            sheet_name = receipt_data['sheet_name']
            
            # Process each image in this sheet
            for img_info in receipt_data.get('images', []):
                img_path = img_info['path']
                
                # Extract date from image if possible
                receipt_date = self.ocr_processor.extract_receipt_date(img_path)
                
                if receipt_date:
                    print(f"\n  [OCR] Processing receipt from {sheet_name}, date: {receipt_date}")
                else:
                    print(f"\n  [OCR] Processing receipt from {sheet_name} (no date detected)")
                    # Try to find nearby date in sheet
                    nearby_date = self._find_nearby_date(receipt_data, img_info)
                    if nearby_date:
                        receipt_date = nearby_date
                        print(f"  [OCR] Using nearby sheet date: {receipt_date}")
                
                # Process the receipt image
                items = self.ocr_processor.process_receipt_image(img_path, receipt_date)
                
                if items:
                    # Store in cache by date
                    cache_key = receipt_date or f"unknown_{sheet_name}_{img_info['index']}"
                    self.receipt_items_cache[cache_key] = {
                        'date': receipt_date,
                        'sheet': sheet_name,
                        'items': items,
                        'image_path': img_path
                    }
                    print(f"  [OCR] Cached {len(items)} items for {cache_key}")
    
    def _find_nearby_date(self, receipt_data: Dict, img_info: Dict) -> Optional[str]:
        """Find date in sheet near the image position"""
        img_row = img_info.get('row', 0)
        
        # Look for dates close to this row
        for date_info in receipt_data.get('dates', []):
            date_row = date_info.get('row', 0)
            if abs(date_row - img_row) < 10:  # Within 10 rows
                return date_info['date']
        
        return None
    
    def _match_transactions_with_receipts(self) -> List[Dict[str, Any]]:
        """Match transactions with receipt items based on date and description"""
        print(f"\n  Matching {len(self.all_transactions)} transactions with receipt items...")
        
        matched_items = []
        unmatched_transactions = []
        
        # Group transactions by date
        transactions_by_date = {}
        for trans in self.all_transactions:
            tgl = trans.get('tanggal')
            if tgl:
                if tgl not in transactions_by_date:
                    transactions_by_date[tgl] = []
                transactions_by_date[tgl].append(trans)
        
        # Match each date group
        for date, transactions in transactions_by_date.items():
            print(f"\n  [MATCH] Date: {date} - {len(transactions)} transactions")
            
            # Find receipt items for this date
            receipt_data = self.receipt_items_cache.get(date)
            
            if receipt_data:
                receipt_items = receipt_data['items']
                print(f"    Found receipt with {len(receipt_items)} items")
                
                # Track which receipt items have been used
                used_receipt_items = set()
                
                # First pass: try to match each transaction with best receipt item
                for trans_idx, trans in enumerate(transactions):
                    matched = self._match_single_transaction(trans, receipt_items, used_receipt_items)
                    if matched:
                        matched_items.append(matched)
                    else:
                        # If no match, save for second pass
                        unmatched_transactions.append({
                            'transaction': trans,
                            'date': date,
                            'index': trans_idx
                        })
                
                # Second pass: assign remaining receipt items to unmatched transactions
                if unmatched_transactions and len(receipt_items) > len(used_receipt_items):
                    print(f"    Second pass: {len(unmatched_transactions)} unmatched transactions, {len(receipt_items) - len(used_receipt_items)} unused receipt items")
                    
                    # Get unused receipt items
                    unused_items = []
                    for i, item in enumerate(receipt_items):
                        if i not in used_receipt_items:
                            unused_items.append(item)
                    
                    # Assign each unused item to an unmatched transaction
                    for i, unmatched in enumerate(unmatched_transactions[:]):
                        if i < len(unused_items):
                            receipt_item = unused_items[i]
                            trans = unmatched['transaction']
                            
                            print(f"    → Assigning receipt item '{receipt_item.get('item_name')}' to transaction")
                            
                            product = {
                                "nama_barang": receipt_item.get('item_name'),
                                "jenis_tipe_barang": self._categorize_item(receipt_item.get('item_name')),
                                "spesifikasi_detail_barang": {
                                    "quantity": receipt_item.get('quantity', 1),
                                    "original_keterangan": trans['keterangan']
                                },
                                "merk_barang": None,
                                "harga_barang": receipt_item.get('item_price') or receipt_item.get('total_price') or trans.get('operasional'),
                                "_page": trans.get('_page', 1),
                                "_date": date,
                                "_matched_with_receipt": True,
                                "_assignment": "forced"
                            }
                            matched_items.append(product)
                            unmatched_transactions.remove(unmatched)
                
                # Third pass: handle remaining unmatched transactions with LLM
                for unmatched in unmatched_transactions:
                    trans = unmatched['transaction']
                    print(f"    → No receipt match, using LLM for transaction")
                    product = self.llm_processor.extract_product_details_llm(
                        trans['keterangan'],
                        trans.get('operasional'),
                        trans.get('_page', 1)
                    )
                    product['_matched_with_receipt'] = False
                    product['_source'] = 'transaction_only'
                    product['_date'] = date
                    matched_items.append(product)
                
                # Clear unmatched for next date
                unmatched_transactions = []
                
            else:
                print(f"    No receipt found for this date")
                # No receipt for this date, use LLM on all transactions
                for trans in transactions:
                    product = self.llm_processor.extract_product_details_llm(
                        trans['keterangan'],
                        trans.get('operasional'),
                        trans.get('_page', 1)
                    )
                    product['_matched_with_receipt'] = False
                    product['_source'] = 'transaction_only'
                    product['_date'] = date
                    matched_items.append(product)
        
        return matched_items
    
    def _match_single_transaction(self, transaction: Dict, receipt_items: List[Dict], used_items: set) -> Optional[Dict]:
        """Match a single transaction with receipt items"""
        keterangan = transaction.get('keterangan', '').lower()
        amount = transaction.get('operasional')
        
        # Split keterangan into individual items
        keterangan_items = self._split_keterangan(keterangan)
        print(f"    Processing transaction: {keterangan[:50]}...")
        print(f"      Split into {len(keterangan_items)} items: {keterangan_items}")
        
        best_match = None
        best_score = 0
        best_item_idx = -1
        best_item_name = ""
        matched_keywords = []
        
        for idx, receipt_item in enumerate(receipt_items):
            if idx in used_items:
                continue
                
            item_name = receipt_item.get('item_name', '').lower()
            item_price = receipt_item.get('item_price') or receipt_item.get('total_price')
            
            # Calculate match score based on multiple factors
            score = 0
            item_matched_keywords = []
            
            # Check for exact matches with any of the keterangan items
            for ket_item in keterangan_items:
                # Exact match
                if ket_item == item_name:
                    score += 20
                    item_matched_keywords.append(f"exact:'{ket_item}'")
                
                # Partial match (one contains the other)
                elif ket_item in item_name or item_name in ket_item:
                    score += 10
                    item_matched_keywords.append(f"partial:'{ket_item}'")
                
                # Check for word matches
                ket_words = set(ket_item.split())
                item_words = set(item_name.split())
                common_words = ket_words.intersection(item_words)
                
                for word in common_words:
                    if len(word) > 2:  # Only count meaningful words
                        score += 3
                        item_matched_keywords.append(f"word:'{word}'")
                
                # Check for BBM/fuel matches
                if 'bbm' in ket_item and ('bbm' in item_name or 'bensin' in item_name):
                    score += 15
                    item_matched_keywords.append("bbm_match")
                
                # Check for specific keywords
                keywords = {
                    'klx': 5, 'vixion': 5, 'revo': 5, 'motor': 3, 'mobil': 3,
                    'galon': 8, 'air': 4, 'minum': 4,
                    'tali': 6, 'rafia': 8,
                    'sapu': 5, 'canebo': 8, 'tisue': 5, 'tisu': 5,
                    'ban': 8, 'ban dalam': 10,
                    'listrik': 6, 'colokan': 7, 'terminal': 7
                }
                
                for kw, kw_score in keywords.items():
                    if kw in ket_item and kw in item_name:
                        score += kw_score
                        item_matched_keywords.append(f"kw:'{kw}'")
            
            # Price proximity check
            if amount and item_price:
                try:
                    price_diff = abs(amount - item_price) / amount if amount > 0 else 1
                    if price_diff < 0.1:  # Within 10%
                        score += 20
                        item_matched_keywords.append(f"price_exact:{amount}")
                    elif price_diff < 0.2:  # Within 20%
                        score += 10
                        item_matched_keywords.append(f"price_close:{amount}")
                    elif price_diff < 0.3:  # Within 30%
                        score += 5
                        item_matched_keywords.append(f"price_near:{amount}")
                except:
                    pass
            
            # Bonus if amount matches exactly
            if amount and item_price and amount == item_price:
                score += 30
                item_matched_keywords.append("price_exact_match")
            
            if score > best_score:
                best_score = score
                best_match = receipt_item
                best_item_idx = idx
                best_item_name = item_name
                matched_keywords = item_matched_keywords
        
        # Minimum threshold for match (reduced to catch more matches)
        if best_match and best_score >= 5:
            print(f"    ✓ Matched with receipt item: {best_match.get('item_name')} (score: {best_score})")
            if matched_keywords:
                print(f"      Matched keywords: {', '.join(matched_keywords[:5])}")
            used_items.add(best_item_idx)
            
            return {
                "nama_barang": best_match.get('item_name'),
                "jenis_tipe_barang": self._categorize_item(best_match.get('item_name')),
                "spesifikasi_detail_barang": {
                    "quantity": best_match.get('quantity', 1),
                    "original_keterangan": transaction['keterangan'],
                    "matched_keywords": matched_keywords[:5]
                },
                "merk_barang": None,
                "harga_barang": best_match.get('item_price') or best_match.get('total_price') or amount,
                "_page": transaction.get('_page', 1),
                "_date": transaction.get('tanggal'),
                "_matched_with_receipt": True,
                "_match_score": best_score
            }
        
        return None

class LLMSemanticProcessor:
    """LLM Semantic Reasoning untuk ekstrak detail produk"""
    
    def __init__(self, api_key: str = None, base_url: str = None):
        """Initialize dengan OpenAI client untuk Qwen"""
        self.api_key = api_key or DASHSCOPE_API_KEY
        self.base_url = base_url or QWEN_BASE_URL
        self.qwen_model = QWEN_MAX_MODEL
        
        self.client = OpenAI(
            api_key=self.api_key,
            base_url=self.base_url
        )
    
    def extract_product_details_llm(self, keterangan_text: str, operasional_amount: float, page_num: int) -> Dict[str, Any]:
        """Use LLM to parse keterangan text and extract product details"""
        if not keterangan_text or keterangan_text.strip() == "":
            return {
                "nama_barang": "Tidak ada keterangan",
                "jenis_tipe_barang": None,
                "spesifikasi_detail_barang": {},
                "merk_barang": None,
                "harga_barang": operasional_amount,
                "_page": page_num
            }
        
        harga_barang = operasional_amount if operasional_amount is not None else None
        
        prompt = f"""Analyze this transaction description and extract product information.

DESCRIPTION: "{keterangan_text}"
AMOUNT: {f'Rp {operasional_amount:,.0f}' if operasional_amount else 'Tidak ada nominal'}

TASK:
Extract the product information into this EXACT JSON structure:
{{
    "nama_barang": "Product name (be specific)",
    "jenis_tipe_barang": "Category like BBM, Alat Tulis, Konsumsi, etc",
    "spesifikasi_detail_barang": {{"detail": "any specifications"}},
    "merk_barang": "Brand if mentioned",
    "harga_barang": {harga_barang if harga_barang is not None else 'null'}
}}

Return ONLY valid JSON, no other text."""
        
        try:
            response = self.client.chat.completions.create(
                model=self.qwen_model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # Extract JSON
            start_idx = result_text.find('{')
            end_idx = result_text.rfind('}') + 1
            
            if start_idx >= 0 and end_idx > start_idx:
                json_str = result_text[start_idx:end_idx]
                product = json.loads(json_str)
            else:
                product = json.loads(result_text)
            
            # Force harga_barang to be actual amount
            product["harga_barang"] = harga_barang
            product["_page"] = page_num
            
            return product
            
        except Exception as e:
            print(f"  ✗ LLM error: {str(e)}")
            
            return {
                "nama_barang": keterangan_text[:100],
                "jenis_tipe_barang": None,
                "spesifikasi_detail_barang": {
                    "deskripsi": keterangan_text
                },
                "merk_barang": None,
                "harga_barang": harga_barang,
                "_page": page_num
            }


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description="Integrated Catalog Extractor - Multi Sheet with OCR from Excel"
    )
    parser.add_argument(
        "excel_file",
        help="Path to Excel file containing cash fund usage data and receipt images"
    )
    parser.add_argument(
        "--output", "-o",
        help="Output JSON file path",
        default=None
    )
    parser.add_argument(
        "--api-key",
        help="DashScope API key",
        default=None
    )
    parser.add_argument(
        "--max-items",
        type=int,
        help="Maximum number of items to process (for testing)",
        default=None
    )
    
    args = parser.parse_args()
    
    # Check if file exists
    if not os.path.exists(args.excel_file):
        print(f"✗ Error: File {args.excel_file} not found")
        return 1
    
    # Initialize integrated extractor
    extractor = IntegratedCatalogExtractor(api_key=args.api_key)
    
    # Process file
    try:
        result = extractor.process_all_sheets(
            args.excel_file,
            args.output
        )
        
    except Exception as e:
        print(f"✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())